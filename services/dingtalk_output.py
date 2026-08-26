"""Template-copy DingTalk output with verified eleven-column backup."""

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import shutil
import time
from collections.abc import Callable, Mapping, Sequence
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import uuid4

from filelock import FileLock

from backend.ai.types import TEST_CASE_FIELDS
from backend.security.redaction import redact_text
from services.dingtalk_mcp import DingTalkMCPService, extract_node_id
from services.dingtalk_spreadsheet import DingTalkSpreadSheetMCPService
from utils.excel_writer import CASE_COLUMNS, ExcelWriter


_FIELD_TO_COLUMN = dict(zip(TEST_CASE_FIELDS, CASE_COLUMNS))
_ROW_PREFIX = re.compile(r"^\[row=\d+\]\s?")
_INVALID_FILENAME = re.compile(r'[\x00-\x1f\x7f-\x9f<>:"/\\|?*]')
_WINDOWS_RESERVED = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}
_FORMULA_AFTER_PREFIX = re.compile(
    r"^[\s\x00-\x1f\x7f-\x9f]*[=+\-@]"
)
_DINGTALK_LITERAL_TRANSLATION = str.maketrans({"\uffe5": "\u00a5"})
_OUTPUT_CLEAR_RANGE = "A:Z"


@dataclass(frozen=True)
class OutputWriteResult:
    dingtalk_doc_url: str | None
    node_id: str | None
    output_file_path: str | None
    partial_failure: bool = False
    local_error: str | None = None


class OutputWriteError(RuntimeError):
    """可安全展示的输出异常，可选保留已经创建的远端节点。"""

    def __init__(
        self,
        node_id: str | None,
        doc_url: str | None,
        redacted_detail: str,
    ) -> None:
        self.node_id = node_id
        self.doc_url = doc_url
        self.redacted_detail = redact_text(redacted_detail)
        super().__init__(self.redacted_detail)


class DingTalkOutputError(OutputWriteError):
    """A safe remote-output failure, retaining a known copied node."""


class LocalOutputError(OutputWriteError):
    """本地模板输出失败，不在错误中暴露用户文件路径。"""


class DingTalkOutputWriter:
    def __init__(
        self,
        document_service: DingTalkMCPService,
        spreadsheet_service: DingTalkSpreadSheetMCPService,
        document_template_url: str,
        output_folder_url: str,
        excel_writer_factory: Callable[[], ExcelWriter] = ExcelWriter,
        poll_interval_seconds: float = 1.0,
        # Large CSV imports are applied asynchronously by DingTalk.  In
        # production a 1,073-row write was correct on readback, but became
        # visible only after the former 300-second window expired.
        # Keep polling bounded while allowing normal large generations to
        # reach their consistent state.
        poll_timeout_seconds: float = 600.0,
        lock_dir: Path = Path("data"),
    ) -> None:
        self.document_service = document_service
        self.spreadsheet_service = spreadsheet_service
        self.document_template_url = document_template_url
        self.output_folder_url = output_folder_url
        self.excel_writer_factory = excel_writer_factory
        self.poll_interval_seconds = max(0.0, float(poll_interval_seconds))
        self.poll_timeout_seconds = max(0.0, float(poll_timeout_seconds))
        self.lock_dir = Path(lock_dir)

    def write(
        self,
        title: str,
        cases: Sequence[Mapping[str, Any]],
        local_output_dir: str | Path,
    ) -> OutputWriteResult:
        safe_cases = self._normalize_cases(cases)
        output_folder_id = extract_node_id(self.output_folder_url)
        template_node_id = extract_node_id(self.document_template_url)
        safe_title = self._safe_output_title(title)
        node_id: str | None = None
        doc_url: str | None = None

        lock_digest = hashlib.sha256(
            output_folder_id.encode("utf-8")
        ).hexdigest()
        lock_path = self.lock_dir / f"dingtalk-output-{lock_digest}.lock"

        try:
            self.lock_dir.mkdir(parents=True, exist_ok=True)
            with FileLock(str(lock_path)):
                before_nodes = self.document_service.list_nodes(
                    output_folder_id
                )
                before_ids = self._node_ids(before_nodes)
                copy_result = self.document_service.copy_document(
                    template_node_id,
                    output_folder_id,
                )
                returned_id = self._copy_node_id(copy_result)
                if returned_id:
                    if returned_id in before_ids:
                        raise DingTalkOutputError(
                            None,
                            None,
                            "模板复制返回了创建前已存在的节点",
                        )
                    node_id = self._poll_for_copy(
                        output_folder_id,
                        before_ids,
                        expected_id=returned_id,
                    )
                else:
                    node_id = self._poll_for_copy(
                        output_folder_id,
                        before_ids,
                    )

                doc_url = self._document_url(node_id)
                self.document_service.rename_document(node_id, safe_title)

                sheets = self.spreadsheet_service.get_all_sheets(node_id)
                if not isinstance(sheets, list) or not sheets:
                    raise RuntimeError("未找到可写工作表")
                sheet_id = sheets[0].get("sheetId")
                if not isinstance(sheet_id, str) or not sheet_id:
                    raise RuntimeError("工作表标识无效")

                csv_text, expected_rows = self._serialize_cases(safe_cases)
                self.spreadsheet_service.clear_range(
                    node_id,
                    sheet_id,
                    _OUTPUT_CLEAR_RANGE,
                    "content",
                )
                self.spreadsheet_service.set_range_from_csv(
                    node_id,
                    sheet_id,
                    "A1",
                    csv_text,
                    True,
                )
                range_address = f"A1:K{len(safe_cases) + 1}"
                self._poll_for_verified_readback(
                    node_id,
                    sheet_id,
                    range_address,
                    expected_rows,
                )
        except DingTalkOutputError:
            raise
        except Exception as exc:
            stage = "钉钉输出"
            raise DingTalkOutputError(
                node_id,
                doc_url,
                f"{stage}失败（{type(exc).__name__}）",
            ) from None

        try:
            output_path = self._local_output_path(
                local_output_dir,
                safe_title,
            )
            self._write_local_backup(output_path, safe_cases)
        except Exception as exc:
            return OutputWriteResult(
                dingtalk_doc_url=doc_url,
                node_id=node_id,
                output_file_path=None,
                partial_failure=True,
                local_error=f"本地备份写入失败（{type(exc).__name__}）",
            )

        return OutputWriteResult(
            dingtalk_doc_url=doc_url,
            node_id=node_id,
            output_file_path=str(output_path),
        )

    def recover_existing(
        self,
        node_id: str,
        expected_case_count: int,
        local_output_dir: str | Path,
    ) -> OutputWriteResult:
        """Re-verify an already written remote result after a transient read."""
        sheets = self.spreadsheet_service.get_all_sheets(node_id)
        if not isinstance(sheets, list) or not sheets:
            raise DingTalkOutputError(node_id, self._document_url(node_id), "未找到可读工作表")
        sheet_id = sheets[0].get("sheetId")
        if not isinstance(sheet_id, str) or not sheet_id:
            raise DingTalkOutputError(node_id, self._document_url(node_id), "工作表标识无效")
        range_address = f"A1:K{max(0, int(expected_case_count)) + 1}"
        try:
            readback = self.spreadsheet_service.get_range_as_csv(
                node_id,
                sheet_id,
                range_address,
            )
            rows = self._verify_readback_shape(
                readback,
                expected_case_count=max(0, int(expected_case_count)),
            )
        except DingTalkOutputError:
            raise
        except Exception as exc:
            raise DingTalkOutputError(
                node_id,
                self._document_url(node_id),
                f"钉钉已有输出重新验收失败（{type(exc).__name__}）",
            ) from None

        # CSV readback exposes displayed values and may omit the apostrophe
        # that protected a formula-like literal on import.  Normalize again
        # before creating an XLSX backup so recovery cannot reintroduce a
        # spreadsheet formula locally.
        safe_cases = self._normalize_cases([
            dict(zip(TEST_CASE_FIELDS, row))
            for row in rows[1:]
        ])
        try:
            title = self.document_service.get_document_name(node_id)
            safe_title = self._safe_output_title(title)
            output_path = self._local_output_path(local_output_dir, safe_title)
            self._write_local_backup(output_path, safe_cases)
        except Exception as exc:
            return OutputWriteResult(
                dingtalk_doc_url=self._document_url(node_id),
                node_id=node_id,
                output_file_path=None,
                partial_failure=True,
                local_error=f"本地备份写入失败（{type(exc).__name__}）",
            )
        return OutputWriteResult(
            dingtalk_doc_url=self._document_url(node_id),
            node_id=node_id,
            output_file_path=str(output_path),
        )

    def _write_local_backup(
        self,
        output_path: Path,
        cases: Sequence[Mapping[str, Any]],
    ) -> None:
        """串行且原子地写本地备份，避免同名并行任务相互截断。"""

        output_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path = output_path.with_suffix(output_path.suffix + ".lock")
        temporary = output_path.with_name(
            f".{output_path.stem}.{uuid4().hex}.tmp{output_path.suffix}"
        )
        with FileLock(str(lock_path)):
            try:
                local_writer = self.excel_writer_factory()
                local_writer.add_test_cases(list(cases))
                local_writer.write_to_excel(str(temporary))
                os.replace(temporary, output_path)
            finally:
                temporary.unlink(missing_ok=True)

    def _poll_for_verified_readback(
        self,
        node_id: str,
        sheet_id: str,
        range_address: str,
        expected_rows: list[list[str]],
    ) -> None:
        deadline = time.monotonic() + max(self.poll_timeout_seconds, 0.001)
        while True:
            try:
                readback = self.spreadsheet_service.get_range_as_csv(
                    node_id,
                    sheet_id,
                    range_address,
                )
                self._verify_readback(readback, expected_rows)
                return
            except Exception:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    raise RuntimeError("钉钉写入回读在等待后仍不一致") from None
                delay = self.poll_interval_seconds or 0.01
                time.sleep(min(delay, remaining))

    def _poll_for_copy(
        self,
        folder_id: str,
        before_ids: set[str],
        expected_id: str | None = None,
    ) -> str:
        deadline = time.monotonic() + max(
            self.poll_timeout_seconds,
            0.001,
        )
        while True:
            nodes = self._list_nodes_with_deadline(folder_id, deadline)
            visible_ids = self._node_ids(nodes)
            candidates = {
                node_id
                for node in nodes
                if self._is_axls(node)
                for node_id in [self._node_id(node)]
                if node_id and node_id not in before_ids
            }
            if expected_id is not None:
                if expected_id in candidates:
                    return expected_id
                if expected_id in visible_ids:
                    raise DingTalkOutputError(
                        None,
                        None,
                        "模板复制返回节点不是新表格",
                    )
            elif len(candidates) == 1:
                return next(iter(candidates))
            if expected_id is None and len(candidates) > 1:
                raise DingTalkOutputError(
                    None,
                    None,
                    "模板复制产生多个新表格节点，拒绝猜测",
                )
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                raise DingTalkOutputError(
                    None,
                    None,
                    "等待模板复制结果超时",
                )
            delay = self.poll_interval_seconds or 0.01
            time.sleep(min(delay, remaining))

    def _list_nodes_with_deadline(
        self,
        folder_id: str,
        deadline: float,
    ) -> list[dict[str, Any]]:
        setter = getattr(self.document_service, "set_request_deadline", None)
        if callable(setter):
            setter(deadline)
        try:
            return self.document_service.list_nodes(folder_id)
        finally:
            if callable(setter):
                setter(None)

    @classmethod
    def _normalize_cases(
        cls,
        cases: Sequence[Mapping[str, Any]],
    ) -> list[dict[str, str]]:
        normalized: list[dict[str, str]] = []
        expected = set(TEST_CASE_FIELDS)
        for case in cases:
            if not isinstance(case, Mapping) or set(case) != expected:
                raise DingTalkOutputError(
                    None,
                    None,
                    "测试用例字段不符合十一列契约",
                )
            row: dict[str, str] = {}
            for field in TEST_CASE_FIELDS:
                value = case[field]
                if value is None:
                    text = ""
                elif isinstance(value, (str, int, float, bool)):
                    text = str(value)
                else:
                    raise DingTalkOutputError(
                        None,
                        None,
                        "测试用例字段值格式无效",
                    )
                text = text.replace("\r\n", "\n").replace("\r", "\n")
                # DingTalk's spreadsheet engine compatibility-folds U+FFE5
                # to U+00A5 on every write path.  Canonicalize before both
                # remote and local output so verification and backups retain
                # one deterministic literal representation.
                text = text.translate(_DINGTALK_LITERAL_TRANSLATION)
                row[field] = cls._neutralize_formula(text)
            normalized.append(row)
        return normalized

    @staticmethod
    def _neutralize_formula(value: str) -> str:
        # Keep CSV and local XLSX semantics identical while preventing formula
        # execution when generated content begins with a spreadsheet sigil.
        # Excel/CSV importers may ignore leading whitespace/control characters,
        # so inspect the first non-whitespace character without removing data.
        formula_risk = _FORMULA_AFTER_PREFIX.match(value) is not None
        excel_safe = "".join(
            f"\\x{codepoint:02x}"
            if (
                codepoint < 32 and character not in {"\t", "\n"}
            )
            or 0x7F <= codepoint <= 0x9F
            else character
            for character in value
            for codepoint in [ord(character)]
        )
        return "'" + excel_safe if formula_risk else excel_safe

    @staticmethod
    def _serialize_cases(
        cases: Sequence[Mapping[str, str]],
    ) -> tuple[str, list[list[str]]]:
        rows = [
            list(CASE_COLUMNS),
            *(
                [case[field] for field in TEST_CASE_FIELDS]
                for case in cases
            ),
        ]
        output = io.StringIO(newline="")
        writer = csv.writer(
            output,
            dialect="excel",
            lineterminator="\r\n",
        )
        writer.writerows(rows)
        return output.getvalue(), rows

    @staticmethod
    def _verify_readback(
        csv_text: str,
        expected_rows: list[list[str]],
    ) -> None:
        actual_rows = DingTalkOutputWriter._parse_readback(csv_text)
        if not actual_rows or actual_rows[0] != list(CASE_COLUMNS):
            raise RuntimeError("回读表头不匹配")
        if len(actual_rows) != len(expected_rows):
            raise RuntimeError("回读行数不匹配")
        if any(len(row) != len(CASE_COLUMNS) for row in actual_rows):
            raise RuntimeError("回读列数不匹配")
        if any(
            not DingTalkOutputWriter._readback_cells_equal(expected, actual)
            for expected_row, actual_row in zip(expected_rows, actual_rows)
            for expected, actual in zip(expected_row, actual_row)
        ):
            raise RuntimeError("回读单元格内容不匹配")

    @staticmethod
    def _readback_cells_equal(expected: str, actual: str) -> bool:
        if expected == actual:
            return True
        # DingTalk protects an imported formula-like literal but omits the
        # leading safety apostrophe when exporting its displayed CSV value.
        # Accept only that exact, narrowly-scoped representation change.  A
        # genuinely evaluated formula would return its result instead of the
        # original sigil-prefixed text and therefore still fail verification.
        return (
            expected.startswith("'")
            and _FORMULA_AFTER_PREFIX.match(expected[1:]) is not None
            and actual == expected[1:]
        )

    @staticmethod
    def _verify_readback_shape(
        csv_text: str,
        expected_case_count: int,
    ) -> list[list[str]]:
        actual_rows = DingTalkOutputWriter._parse_readback(csv_text)
        if not actual_rows or actual_rows[0] != list(CASE_COLUMNS):
            raise RuntimeError("回读表头不匹配")
        if len(actual_rows) != expected_case_count + 1:
            raise RuntimeError("回读行数不匹配")
        if any(len(row) != len(CASE_COLUMNS) for row in actual_rows):
            raise RuntimeError("回读列数不匹配")
        return actual_rows

    @staticmethod
    def _parse_readback(csv_text: str) -> list[list[str]]:
        if not isinstance(csv_text, str):
            raise RuntimeError("回读未返回 CSV 文本")
        unprefixed = "".join(
            _ROW_PREFIX.sub("", line)
            for line in csv_text.splitlines(keepends=True)
        )
        try:
            actual_rows = list(csv.reader(io.StringIO(unprefixed)))
        except csv.Error as exc:
            raise RuntimeError("回读 CSV 无效") from exc
        return actual_rows

    @staticmethod
    def _safe_output_title(title: str) -> str:
        text = _INVALID_FILENAME.sub("_", str(title)).strip().rstrip(". ")
        text = re.sub(r"(?:-用例)+$", "-用例", text)
        if not text:
            return "测试用例"
        if text.upper() in _WINDOWS_RESERVED:
            text = f"_{text}"
        if text.endswith("-用例") or text == "测试用例":
            suffix = "-用例" if text.endswith("-用例") else ""
            base = text[: -len(suffix)] if suffix else text
        else:
            base = text
            suffix = "-用例"
        maximum_base = 100 - len(suffix)
        base = base[:maximum_base].rstrip(". ") or "测试"
        return base + suffix

    @staticmethod
    def _local_output_path(
        local_output_dir: str | Path,
        safe_title: str,
    ) -> Path:
        root = Path(local_output_dir).expanduser().resolve()
        destination = (root / f"{safe_title}.xlsx").resolve()
        if destination.parent != root:
            raise DingTalkOutputError(
                None,
                None,
                "本地输出路径越界",
            )
        return destination

    @staticmethod
    def _node_id(node: Mapping[str, Any]) -> str | None:
        for key in ("nodeId", "id", "dentryUuid"):
            value = node.get(key)
            if isinstance(value, str) and value:
                return value
        return None

    @classmethod
    def _node_ids(cls, nodes: object) -> set[str]:
        if not isinstance(nodes, list):
            raise RuntimeError("输出目录节点列表无效")
        return {
            node_id
            for node in nodes
            if isinstance(node, Mapping)
            for node_id in [cls._node_id(node)]
            if node_id
        }

    @classmethod
    def _copy_node_id(cls, result: object) -> str | None:
        if not isinstance(result, Mapping):
            raise RuntimeError("模板复制结果无效")
        node_id = cls._node_id(result)
        if node_id is None:
            return None
        return extract_node_id(node_id)

    @staticmethod
    def _is_axls(node: Mapping[str, Any]) -> bool:
        value = node.get(
            "extension",
            node.get("type", node.get("fileType", node.get("nodeType"))),
        )
        return isinstance(value, str) and value.lower().lstrip(".") == "axls"

    @staticmethod
    def _document_url(node_id: str) -> str:
        return f"https://alidocs.dingtalk.com/i/nodes/{node_id}"


class LocalTemplateOutputWriter:
    """在用户默认模板副本上生成本地结果，不依赖钉钉输出服务。"""

    def __init__(self, template_path: str | Path) -> None:
        self.template_path = Path(template_path).expanduser().resolve()

    def write(
        self,
        title: str,
        cases: Sequence[Mapping[str, Any]],
        local_output_dir: str | Path,
    ) -> OutputWriteResult:
        safe_cases = DingTalkOutputWriter._normalize_cases(cases)
        safe_title = DingTalkOutputWriter._safe_output_title(title)
        output_path = DingTalkOutputWriter._local_output_path(
            local_output_dir,
            safe_title,
        )
        if not self.template_path.is_file():
            raise LocalOutputError(None, None, "本地输出模板不可用")
        if output_path == self.template_path:
            raise LocalOutputError(None, None, "本地输出路径不能覆盖模板")
        try:
            self._write_from_template(output_path, safe_cases)
        except OutputWriteError:
            raise
        except Exception as exc:
            raise LocalOutputError(
                None,
                None,
                f"本地模板输出失败（{type(exc).__name__}）",
            ) from None
        return OutputWriteResult(
            dingtalk_doc_url=None,
            node_id=None,
            output_file_path=str(output_path),
        )

    def _write_from_template(
        self,
        output_path: Path,
        cases: Sequence[Mapping[str, str]],
    ) -> None:
        """复制完整工作簿后原子写入首个十一列表，保留其它说明工作表。"""

        from openpyxl import load_workbook

        output_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path = output_path.with_suffix(output_path.suffix + ".lock")
        temporary = output_path.with_name(
            f".{output_path.stem}.{uuid4().hex}.tmp{output_path.suffix}"
        )
        with FileLock(str(lock_path)):
            try:
                shutil.copyfile(self.template_path, temporary)
                workbook = load_workbook(temporary)
                try:
                    sheet = self._case_sheet(workbook.worksheets)
                    original_max_row = max(2, int(sheet.max_row or 2))
                    target_max_row = max(original_max_row, len(cases) + 1)
                    for row_index in range(2, target_max_row + 1):
                        for column_index in range(1, len(CASE_COLUMNS) + 1):
                            sheet.cell(row_index, column_index).value = None

                    for row_index, case in enumerate(cases, start=2):
                        if row_index > original_max_row:
                            self._copy_row_style(sheet, 2, row_index)
                        for column_index, field in enumerate(
                            TEST_CASE_FIELDS,
                            start=1,
                        ):
                            sheet.cell(row_index, column_index).value = case[field]
                    workbook.save(temporary)
                finally:
                    workbook.close()
                os.replace(temporary, output_path)
            finally:
                temporary.unlink(missing_ok=True)

    @staticmethod
    def _case_sheet(worksheets):
        expected = list(CASE_COLUMNS)
        for sheet in worksheets:
            actual = [
                str(sheet.cell(1, column).value or "").strip()
                for column in range(1, len(expected) + 1)
            ]
            if actual == expected:
                return sheet
        raise LocalOutputError(None, None, "本地输出模板表头不符合十一列契约")

    @staticmethod
    def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
        """超出预置空白行时复制样式，避免大批量用例突然失去格式。"""

        if source_row in sheet.row_dimensions:
            sheet.row_dimensions[target_row].height = sheet.row_dimensions[
                source_row
            ].height
        for column in range(1, len(CASE_COLUMNS) + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format


__all__ = [
    "CASE_COLUMNS",
    "DingTalkOutputError",
    "DingTalkOutputWriter",
    "LocalOutputError",
    "LocalTemplateOutputWriter",
    "OutputWriteResult",
    "OutputWriteError",
]
