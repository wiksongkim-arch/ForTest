"""默认模板母版与用户工作副本测试。"""

from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from services.dingtalk_output import LocalTemplateOutputWriter
from utils.template_loader import TemplateLoader
from utils.default_templates import (
    CONTENT_TEMPLATE,
    OUTPUT_TEMPLATE,
    DefaultTemplateManager,
    bundled_templates_root,
)


def _digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_default_templates_are_valid_xlsx_assets_and_create_user_copies(
    tmp_path: Path,
):
    manager = DefaultTemplateManager(tmp_path)

    paths = manager.ensure_all()

    assert set(paths) == {CONTENT_TEMPLATE, OUTPUT_TEMPLATE}
    assert all(path.parent == tmp_path / "templates" for path in paths.values())
    for path in paths.values():
        workbook = load_workbook(path, read_only=True)
        try:
            assert workbook.sheetnames
        finally:
            workbook.close()


def test_existing_user_copy_is_preserved_until_explicit_restore(tmp_path: Path):
    manager = DefaultTemplateManager(tmp_path)
    user_copy = manager.ensure(CONTENT_TEMPLATE)
    bundled = bundled_templates_root() / "test_case_content_template.xlsx"
    bundled_digest = _digest(bundled)
    user_copy.write_bytes(b"user-customized-template")

    assert manager.ensure(CONTENT_TEMPLATE).read_bytes() == b"user-customized-template"

    restored = manager.restore(CONTENT_TEMPLATE)

    assert _digest(restored) == bundled_digest
    # 恢复只读取程序母版，母版本身的内容始终保持不变。
    assert _digest(bundled) == bundled_digest


def test_content_template_loader_reads_user_xlsx_without_online_service(
    tmp_path: Path,
):
    content_path = DefaultTemplateManager(tmp_path).ensure(CONTENT_TEMPLATE)

    template = TemplateLoader(
        "",
        None,
        local_template_path=content_path,
    ).load_template(force_refresh=True)

    assert template["field_specs"]["用例名称"]
    assert "筛选-搜索" in template["components"]
    assert template["components"]["筛选-搜索"]


def test_local_output_uses_template_and_preserves_instruction_sheets(tmp_path: Path):
    output_template = DefaultTemplateManager(tmp_path).ensure(OUTPUT_TEMPLATE)
    source_digest = _digest(output_template)
    output_root = tmp_path / "output"
    case = {
        "module": "登录",
        "case_name": "账号密码登录",
        "prerequisite": "已有账号",
        "test_steps": "1. 输入账号密码",
        "expected_result": "登录成功",
        "priority": "高",
        "case_type": "功能测试",
        "applicable_phase": "系统测试",
        "remark": "",
        "case_id": "TC-001",
        "execution": "未执行",
    }

    result = LocalTemplateOutputWriter(output_template).write(
        "登录需求",
        [case],
        output_root,
    )

    assert result.dingtalk_doc_url is None
    assert result.node_id is None
    output_path = Path(str(result.output_file_path))
    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook.sheetnames == [
            "用例按功能模块拆分sheet",
            "用例按功能模块拆分sheet-初版",
            "埋点",
            "权限",
            "其他",
            "用例字段说明",
        ]
        row = [
            workbook.worksheets[0].cell(2, column).value
            for column in range(1, 12)
        ]
        expected_row = list(case.values())
        expected_row[8] = None  # Excel 将空字符串规范化为空单元格。
        assert row == expected_row
    finally:
        workbook.close()
    assert _digest(output_template) == source_digest
