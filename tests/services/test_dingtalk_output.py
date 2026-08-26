import csv
import hashlib
import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl

from backend.ai.types import TEST_CASE_FIELDS
from services.dingtalk_mcp import DingTalkMCPError
from services.dingtalk_output import (
    CASE_COLUMNS,
    DingTalkOutputError,
    DingTalkOutputWriter,
)
from services.dingtalk_spreadsheet import DingTalkSpreadSheetMCPService
from utils.excel_writer import ExcelWriter


def make_case(**overrides):
    case = dict(zip(TEST_CASE_FIELDS, ["value"] * len(TEST_CASE_FIELDS)))
    case.update(overrides)
    return case


def configure_confirmed_copy(document, node_id="new-node"):
    document.list_nodes.side_effect = (
        [],
        [{"nodeId": node_id, "extension": "axls"}],
    )
    document.copy_document.return_value = {"nodeId": node_id}


class EchoSpreadsheetService:
    def __init__(self, *, readback=None):
        self.calls = []
        self.csv_text = ""
        self.readback = readback

    def get_all_sheets(self, node_id):
        self.calls.append(("sheets", node_id))
        return [{"sheetId": "sheet-1"}]

    def clear_range(self, node_id, sheet_id, range_address, clear_type):
        self.calls.append(("clear", range_address, clear_type))

    def set_range_from_csv(
        self, node_id, sheet_id, start_cell, csv_text, allow_overwrite
    ):
        self.calls.append(("set", start_cell, allow_overwrite))
        self.csv_text = csv_text

    def get_range_as_csv(self, node_id, sheet_id, range_address):
        self.calls.append(("read", range_address))
        if callable(self.readback):
            return self.readback(self.csv_text)
        return self.csv_text if self.readback is None else self.readback


class DingTalkOutputTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)

    def tearDown(self):
        self.temp.cleanup()

    def writer(self, document, sheets, local=None, **overrides):
        if local is None:
            local = Mock()
            # 原子写实现要求写入器确实创建临时文件。
            local.write_to_excel.side_effect = lambda path: Path(path).touch()
        values = {
            "document_service": document,
            "spreadsheet_service": sheets,
            "document_template_url": (
                "https://alidocs.dingtalk.com/i/nodes/template-node"
            ),
            "output_folder_url": (
                "https://alidocs.dingtalk.com/i/nodes/folder-node"
            ),
            "excel_writer_factory": Mock(return_value=local),
            "poll_interval_seconds": 0,
            "poll_timeout_seconds": 0.05,
            "lock_dir": self.root / "locks",
        }
        values.update(overrides)
        return DingTalkOutputWriter(**values), local

    def test_default_poll_window_allows_large_remote_imports_to_settle(self):
        writer = DingTalkOutputWriter(
            document_service=Mock(),
            spreadsheet_service=Mock(),
            document_template_url=(
                "https://alidocs.dingtalk.com/i/nodes/template-node"
            ),
            output_folder_url=(
                "https://alidocs.dingtalk.com/i/nodes/folder-node"
            ),
        )

        self.assertEqual(writer.poll_timeout_seconds, 600.0)

    def test_columns_are_the_single_approved_eleven_column_contract(self):
        self.assertEqual(
            CASE_COLUMNS,
            [
                "所属模块",
                "用例名称",
                "前置条件",
                "测试步骤",
                "预期结果",
                "优先级",
                "用例类型",
                "适用阶段",
                "备注",
                "用例编号",
                "执行状态",
            ],
        )
        self.assertEqual(ExcelWriter.COLUMNS, CASE_COLUMNS)

    def test_writer_copies_template_clears_writes_and_verifies_all_cells(self):
        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService()
        writer, local = self.writer(document, sheets)
        case = make_case(case_name='contains, comma and "quote"')

        with patch("services.dingtalk_output.FileLock") as file_lock:
            result = writer.write("登录需求", [case], self.root / "output")

        document.copy_document.assert_called_once_with(
            "template-node", "folder-node"
        )
        document.rename_document.assert_called_once_with(
            "new-node", "登录需求-用例"
        )
        self.assertEqual(
            sheets.calls,
            [
                ("sheets", "new-node"),
                ("clear", "A:Z", "content"),
                ("set", "A1", True),
                ("read", "A1:K2"),
            ],
        )
        rows = list(csv.reader(io.StringIO(sheets.csv_text)))
        self.assertEqual(rows[0], CASE_COLUMNS)
        self.assertEqual(rows[1][1], case["case_name"])
        self.assertEqual(len(rows[1]), 11)
        local.add_test_cases.assert_called_once_with([case])
        temporary_path = Path(local.write_to_excel.call_args.args[0])
        output_path = Path(result.output_file_path)
        self.assertEqual(temporary_path.parent, (self.root / "output").resolve())
        self.assertEqual(output_path.name, "登录需求-用例.xlsx")
        self.assertEqual(result.node_id, "new-node")
        self.assertFalse(result.partial_failure)
        expected_lock = (
            self.root
            / "locks"
            / (
                "dingtalk-output-"
                + hashlib.sha256(b"folder-node").hexdigest()
                + ".lock"
            )
        )
        file_lock.assert_any_call(str(expected_lock))
        file_lock.assert_any_call(str(output_path.with_suffix(".xlsx.lock")))
        document.create_file.assert_not_called()

    def test_copy_without_returned_id_accepts_only_one_new_axls_node(self):
        document = Mock()
        node_lists = iter(
            (
                [{"nodeId": "old", "type": "axls"}],
                [
                    {"nodeId": "old", "extension": "axls"},
                    {"nodeId": "new-doc", "type": "doc"},
                    {"nodeId": "copied", "extension": "axls"},
                ],
            )
        )
        document.list_nodes.side_effect = lambda _folder: next(node_lists)
        document.copy_document.return_value = {}
        writer, _local = self.writer(document, EchoSpreadsheetService())

        result = writer.write("需求", [make_case()], self.root / "out")

        self.assertEqual(result.node_id, "copied")
        document.rename_document.assert_called_once_with(
            "copied", "需求-用例"
        )

    def test_returned_copy_id_is_rejected_when_it_existed_before_copy(self):
        document = Mock()
        document.list_nodes.return_value = [
            {"nodeId": "existing", "type": "axls"}
        ]
        document.copy_document.return_value = {"nodeId": "existing"}
        writer, _local = self.writer(document, EchoSpreadsheetService())

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertIsNone(raised.exception.node_id)
        document.rename_document.assert_not_called()

    def test_returned_copy_id_must_be_confirmed_as_the_new_axls_node(self):
        document = Mock()
        calls = 0

        def list_nodes(_folder):
            nonlocal calls
            calls += 1
            if calls == 1:
                return []
            return [{"nodeId": "different-node", "extension": "axls"}]

        document.list_nodes.side_effect = list_nodes
        document.copy_document.return_value = {"nodeId": "unconfirmed-node"}
        writer, _local = self.writer(
            document,
            EchoSpreadsheetService(),
            poll_timeout_seconds=0,
        )

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertIsNone(raised.exception.node_id)
        document.rename_document.assert_not_called()

    def test_poll_timeout_never_uses_an_unconfirmed_node(self):
        document = Mock()
        document.list_nodes.return_value = []
        document.copy_document.return_value = {}
        writer, _local = self.writer(
            document,
            EchoSpreadsheetService(),
            poll_timeout_seconds=0,
        )

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertIsNone(raised.exception.node_id)
        document.rename_document.assert_not_called()

    def test_copy_polling_never_guesses_between_multiple_new_axls_nodes(self):
        document = Mock()
        calls = 0

        def list_nodes(_folder):
            nonlocal calls
            calls += 1
            if calls == 1:
                return []
            return [
                {"nodeId": "candidate-1", "type": "axls"},
                {"nodeId": "candidate-2", "type": "axls"},
            ]

        document.list_nodes.side_effect = list_nodes
        document.copy_document.return_value = {}
        writer, _local = self.writer(document, EchoSpreadsheetService())

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertIsNone(raised.exception.node_id)
        document.rename_document.assert_not_called()

    def test_known_node_failure_is_recoverable_and_never_deleted(self):
        document = Mock()
        configure_confirmed_copy(document)
        document.rename_document.side_effect = RuntimeError(
            "token=top-secret"
        )
        writer, _local = self.writer(document, EchoSpreadsheetService())

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertEqual(raised.exception.node_id, "new-node")
        self.assertTrue(raised.exception.doc_url.endswith("/new-node"))
        self.assertNotIn("top-secret", str(raised.exception))
        document.delete_document.assert_not_called()

    def test_local_failure_returns_partial_result_and_keeps_remote_node(self):
        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService()
        local = Mock()
        local.write_to_excel.side_effect = OSError("secret disk path")
        writer, _ = self.writer(document, sheets, local=local)

        result = writer.write("需求", [make_case()], self.root / "out")

        self.assertTrue(result.partial_failure)
        self.assertEqual(result.node_id, "new-node")
        self.assertTrue(result.dingtalk_doc_url.endswith("/new-node"))
        self.assertIsNone(result.output_file_path)
        self.assertNotIn("secret disk path", result.local_error)
        document.delete_document.assert_not_called()

    def test_local_path_resolution_failure_is_also_a_partial_result(self):
        document = Mock()
        configure_confirmed_copy(document)
        writer, _local = self.writer(document, EchoSpreadsheetService())

        with patch.object(
            writer,
            "_local_output_path",
            side_effect=OSError("private local path"),
        ):
            result = writer.write("需求", [make_case()], self.root / "out")

        self.assertTrue(result.partial_failure)
        self.assertEqual(result.node_id, "new-node")
        self.assertNotIn("private local path", result.local_error)

    def test_lock_setup_failure_is_typed_and_makes_no_remote_call(self):
        document = Mock()
        blocking_file = self.root / "not-a-directory"
        blocking_file.write_text("block", encoding="utf-8")
        writer, _local = self.writer(
            document,
            EchoSpreadsheetService(),
            lock_dir=blocking_file / "locks",
        )

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertIsNone(raised.exception.node_id)
        document.list_nodes.assert_not_called()
        document.copy_document.assert_not_called()

    def test_title_and_filename_are_sanitized_with_idempotent_suffix(self):
        document = Mock()
        configure_confirmed_copy(document)
        writer, local = self.writer(document, EchoSpreadsheetService())

        result = writer.write(
            "../坏:名称\x00-用例-用例. ",
            [make_case()],
            self.root / "strict-root",
        )

        renamed = document.rename_document.call_args.args[1]
        self.assertEqual(renamed.count("-用例"), 1)
        self.assertNotRegex(renamed, r'[\x00-\x1f<>:"/\\|?*]')
        output_path = Path(result.output_file_path)
        self.assertEqual(
            output_path.parent,
            (self.root / "strict-root").resolve(),
        )

    def test_title_removes_c0_del_and_c1_from_remote_and_local_names(self):
        document = Mock()
        configure_confirmed_copy(document)
        writer, local = self.writer(document, EchoSpreadsheetService())

        result = writer.write(
            "控\x00制\x1f字\x7f符\x85标\x9f题",
            [make_case()],
            self.root / "strict-root",
        )

        renamed = document.rename_document.call_args.args[1]
        output_path = Path(result.output_file_path)
        self.assertEqual(output_path.stem, renamed)
        self.assertNotRegex(renamed, r"[\x00-\x1f\x7f-\x9f]")
        self.assertNotRegex(output_path.name, r"[\x00-\x1f\x7f-\x9f]")

    def test_polling_propagates_and_then_clears_absolute_request_deadline(self):
        class DeadlineDocument:
            def __init__(self):
                self.deadlines = []
                self.list_count = 0

            def set_request_deadline(self, deadline):
                self.deadlines.append(deadline)

            def list_nodes(self, _folder):
                self.list_count += 1
                if self.list_count == 1:
                    return []
                return [{"nodeId": "new-node", "extension": "axls"}]

            def copy_document(self, _node, _folder):
                return {}

            def rename_document(self, _node, _name):
                return None

        document = DeadlineDocument()
        writer, _local = self.writer(document, EchoSpreadsheetService())

        writer.write("需求", [make_case()], self.root / "out")

        finite = [value for value in document.deadlines if value is not None]
        self.assertTrue(finite)
        self.assertEqual(document.deadlines[-1], None)

    def test_row_prefixed_readback_is_accepted(self):
        def prefix(csv_text):
            return "\n".join(
                f"[row={index}] {line}"
                for index, line in enumerate(csv_text.splitlines(), start=1)
            )

        document = Mock()
        configure_confirmed_copy(document)
        writer, _local = self.writer(
            document, EchoSpreadsheetService(readback=prefix)
        )

        result = writer.write("需求", [make_case()], self.root / "out")

        self.assertFalse(result.partial_failure)

    def test_transient_readback_mismatch_is_polled_until_consistent(self):
        attempts = 0

        def eventually_consistent(csv_text):
            nonlocal attempts
            attempts += 1
            if attempts == 1:
                return csv_text.splitlines()[0] + "\r\n"
            return csv_text

        document = Mock()
        configure_confirmed_copy(document)
        writer, _local = self.writer(
            document,
            EchoSpreadsheetService(readback=eventually_consistent),
        )

        result = writer.write("需求", [make_case()], self.root / "out")

        self.assertFalse(result.partial_failure)
        self.assertGreaterEqual(attempts, 2)

    def test_readback_value_mismatch_is_a_typed_recoverable_error(self):
        document = Mock()
        configure_confirmed_copy(document)
        wrong = io.StringIO(newline="")
        writer_csv = csv.writer(wrong)
        writer_csv.writerow(CASE_COLUMNS)
        writer_csv.writerow(["wrong"] * len(CASE_COLUMNS))
        writer, local = self.writer(
            document, EchoSpreadsheetService(readback=wrong.getvalue())
        )

        with self.assertRaises(DingTalkOutputError) as raised:
            writer.write("需求", [make_case()], self.root / "out")

        self.assertEqual(raised.exception.node_id, "new-node")
        local.write_to_excel.assert_not_called()

    def test_readback_rejects_header_row_count_and_column_count_mismatch(self):
        valid_header = io.StringIO(newline="")
        csv.writer(valid_header).writerow(CASE_COLUMNS)
        short_row = io.StringIO(newline="")
        short_writer = csv.writer(short_row)
        short_writer.writerow(CASE_COLUMNS)
        short_writer.writerow(["value"] * 10)
        bad_header = io.StringIO(newline="")
        bad_writer = csv.writer(bad_header)
        bad_writer.writerow(["wrong"] * 11)
        bad_writer.writerow(["value"] * 11)

        for readback in (
            valid_header.getvalue(),
            short_row.getvalue(),
            bad_header.getvalue(),
        ):
            with self.subTest(readback=readback):
                document = Mock()
                configure_confirmed_copy(document)
                writer, local = self.writer(
                    document,
                    EchoSpreadsheetService(readback=readback),
                )
                with self.assertRaises(DingTalkOutputError) as raised:
                    writer.write("需求", [make_case()], self.root / "out")
                self.assertEqual(raised.exception.node_id, "new-node")
                local.write_to_excel.assert_not_called()

    def test_formula_values_are_neutralized_identically_remotely_and_locally(self):
        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService()
        writer, local = self.writer(document, sheets)

        writer.write(
            "需求",
            [make_case(case_name="=1+1", remark="@SUM(A1:A2)")],
            self.root / "out",
        )

        remote_case = list(csv.reader(io.StringIO(sheets.csv_text)))[1]
        local_case = local.add_test_cases.call_args.args[0][0]
        self.assertEqual(remote_case[1], "'=1+1")
        self.assertEqual(remote_case[8], "'@SUM(A1:A2)")
        self.assertEqual(local_case["case_name"], remote_case[1])
        self.assertEqual(local_case["remark"], remote_case[8])

    def test_readback_accepts_dingtalk_display_value_without_safety_apostrophe(self):
        def displayed_csv(csv_text):
            rows = list(csv.reader(io.StringIO(csv_text)))
            rows[1][1] = rows[1][1][1:]
            output = io.StringIO(newline="")
            csv.writer(output, lineterminator="\r\n").writerows(rows)
            return output.getvalue()

        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService(readback=displayed_csv)
        writer, _local = self.writer(document, sheets)

        result = writer.write(
            "需求",
            [make_case(case_name="=1+1")],
            self.root / "out",
        )

        self.assertFalse(result.partial_failure)

    def test_readback_still_rejects_missing_apostrophe_for_non_formula_text(self):
        expected = [list(CASE_COLUMNS), ["'plain", *("value" for _ in range(10))]]
        actual = io.StringIO(newline="")
        csv.writer(actual, lineterminator="\r\n").writerows(
            [list(CASE_COLUMNS), ["plain", *("value" for _ in range(10))]]
        )

        with self.assertRaisesRegex(RuntimeError, "回读单元格内容不匹配"):
            DingTalkOutputWriter._verify_readback(actual.getvalue(), expected)

    def test_dingtalk_compatibility_sign_is_canonicalized_remotely_and_locally(self):
        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService()
        writer, local = self.writer(document, sheets)

        result = writer.write(
            "需求",
            [make_case(test_steps="输入特殊字符￥")],
            self.root / "out",
        )

        self.assertFalse(result.partial_failure)
        remote = list(csv.reader(io.StringIO(sheets.csv_text)))[1][3]
        local_case = local.add_test_cases.call_args.args[0][0]
        self.assertEqual(remote, "输入特殊字符¥")
        self.assertEqual(local_case["test_steps"], remote)

    def test_recovery_restores_formula_safety_prefix_before_local_backup(self):
        document = Mock()
        document.get_document_name.return_value = "恢复用例"
        displayed = io.StringIO(newline="")
        csv.writer(displayed, lineterminator="\r\n").writerows(
            [
                list(CASE_COLUMNS),
                ["value", "=1+1", *("value" for _ in range(9))],
            ]
        )
        sheets = EchoSpreadsheetService(readback=displayed.getvalue())
        writer, local = self.writer(document, sheets)

        result = writer.recover_existing("existing-node", 1, self.root / "out")

        recovered = local.add_test_cases.call_args.args[0][0]
        self.assertEqual(recovered["case_name"], "'=1+1")
        self.assertEqual(result.node_id, "existing-node")

    def test_all_formula_sigils_and_leading_whitespace_are_safe_in_csv_and_xlsx(self):
        dangerous = [
            "=1+1",
            "+1+1",
            "-1+1",
            "@SUM(A1:A2)",
            "\t=1+1",
            "\r+1+1",
            "\n-1+1",
            "  @SUM(A1:A2)",
            "\x00=1+1",
            "\x1f@SUM(A1:A2)",
            "\x85+1+1",
        ]
        document = Mock()
        configure_confirmed_copy(document)
        sheets = EchoSpreadsheetService()
        writer, _local = self.writer(
            document,
            sheets,
            local=ExcelWriter(),
        )

        result = writer.write(
            "公式安全",
            [make_case(case_name=value) for value in dangerous],
            self.root / "out",
        )

        remote_values = [
            row[1]
            for row in list(csv.reader(io.StringIO(sheets.csv_text)))[1:]
        ]
        workbook = openpyxl.load_workbook(
            result.output_file_path,
            read_only=True,
            data_only=False,
        )
        try:
            local_values = [
                row[1]
                for row in list(workbook.active.iter_rows(values_only=True))[1:]
            ]
        finally:
            workbook.close()
        def expected_value(value):
            value = value.replace("\r\n", "\n").replace("\r", "\n")
            value = "".join(
                f"\\x{ord(char):02x}"
                if (
                    ord(char) < 32 and char not in {"\t", "\n"}
                )
                or 0x7F <= ord(char) <= 0x9F
                else char
                for char in value
            )
            return "'" + value

        expected = [expected_value(value) for value in dangerous]
        self.assertEqual(remote_values, expected)
        self.assertEqual(local_values, expected)

    def test_excel_writer_round_trips_all_eleven_columns(self):
        destination = self.root / "real.xlsx"
        case = make_case(
            module="模块",
            case_name="名称",
            case_id="TC-001",
            execution="未执行",
        )
        writer = ExcelWriter()
        writer.add_test_cases([case])
        writer.write_to_excel(str(destination))

        workbook = openpyxl.load_workbook(destination, read_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual(list(rows[0]), CASE_COLUMNS)
        self.assertEqual(len(rows[1]), 11)
        self.assertEqual(rows[1][9], "TC-001")
        self.assertEqual(rows[1][10], "未执行")


class SpreadsheetOutputWrapperTests(unittest.TestCase):
    def test_clear_and_read_csv_wrappers_use_exact_transport_contract(self):
        service = DingTalkSpreadSheetMCPService(
            "https://mcp.example.test/sheets"
        )
        service.call_tool = Mock(
            side_effect=(
                {"structuredContent": {"success": True}},
                {"structuredContent": {"csv": "a,b\r\n1,2\r\n"}},
            )
        )

        service.clear_range("node", "sheet", "A:K", "content")
        csv_text = service.get_range_as_csv("node", "sheet", "A1:K2")

        self.assertEqual(csv_text, "a,b\r\n1,2\r\n")
        self.assertEqual(
            service.call_tool.call_args_list,
            [
                unittest.mock.call(
                    "clear_range",
                    {
                        "nodeId": "node",
                        "sheetId": "sheet",
                        "range": "A:K",
                        "type": "content",
                    },
                ),
                unittest.mock.call(
                    "get_range_as_csv",
                    {
                        "nodeId": "node",
                        "sheetId": "sheet",
                        "range": "A1:K2",
                        "maxChars": 2_000_000,
                    },
                ),
            ],
        )

    def test_read_csv_rejects_truncation_and_missing_text(self):
        for payload in ({"hasMore": True, "csv": "x"}, {"success": True}):
            with self.subTest(payload=payload):
                service = DingTalkSpreadSheetMCPService(
                    "https://mcp.example.test/sheets"
                )
                service.call_tool = Mock(
                    return_value={"structuredContent": payload}
                )
                with self.assertRaises(DingTalkMCPError):
                    service.get_range_as_csv("node", "sheet", "A1:K2")


if __name__ == "__main__":
    unittest.main()
