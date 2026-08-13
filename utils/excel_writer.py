"""
Excel 写入工具 - 将测试用例写入 Excel 文件
"""
import os
from typing import Dict, Any, List
from datetime import datetime
import shutil


CASE_COLUMNS = [
    "所属模块",
    "用例名称",
    "前置条件",
    "测试步骤",
    "预期结果",
    "优先级",
    "用例类型",
    "适用阶段",
    "备注",
    "用例编号",
    "执行状态",
]


class ExcelWriter:
    """Excel 写入器"""

    # Excel 列定义
    COLUMNS = CASE_COLUMNS

    def __init__(self):
        self.test_cases: List[Dict[str, Any]] = []

    def add_test_case(self, case: Dict[str, Any]):
        """添加测试用例"""
        self.test_cases.append(case)

    def add_test_cases(self, cases: List[Dict[str, Any]]):
        """批量添加测试用例"""
        self.test_cases.extend(cases)

    def write_to_excel(self, file_path: str, sheet_name: str = "测试用例"):
        """
        将测试用例写入 Excel 文件

        Args:
            file_path: 文件路径
            sheet_name: 工作表名称
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise Exception("请先安装 openpyxl: pip install openpyxl")

        # 确保目录存在
        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 设置表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 写入表头
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入数据
        for row_idx, case in enumerate(self.test_cases, start=2):
            for col_idx, column_key in enumerate(self.COLUMNS, start=1):
                # 将驼峰命名转换为下划线命名来匹配 case 字典的 key
                key = self._column_to_key(column_key)
                value = case.get(key, "")
                # 处理列表类型的值，转换为换行分隔的字符串
                if isinstance(value, list):
                    value = "\n".join(str(v) for v in value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        # 设置列宽
        column_widths = {
            "所属模块": 20,
            "用例名称": 30,
            "前置条件": 40,
            "测试步骤": 50,
            "预期结果": 40,
            "优先级": 10,
            "用例类型": 15,
            "适用阶段": 15,
            "备注": 30,
            "用例编号": 15,
            "执行状态": 15,
        }
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            width = column_widths.get(column_name, 20)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 保存文件
        wb.save(file_path)

    def _column_to_key(self, column_name: str) -> str:
        """将列名转换为 case 字典的 key"""
        # 简单的映射关系
        mapping = {
            "所属模块": "module",
            "用例名称": "case_name",
            "前置条件": "prerequisite",
            "测试步骤": "test_steps",
            "预期结果": "expected_result",
            "优先级": "priority",
            "用例类型": "case_type",
            "适用阶段": "applicable_phase",
            "备注": "remark",
            "用例编号": "case_id",
            "执行状态": "execution",
        }
        return mapping.get(column_name, column_name)


def copy_template_and_write(
    template_excel_path: str,
    output_path: str,
    test_cases: List[Dict[str, Any]]
) -> str:
    """
    复制 Excel 模板并写入测试用例

    Args:
        template_excel_path: 模板 Excel 文件路径
        output_path: 输出文件路径
        test_cases: 测试用例列表

    Returns:
        生成的文件路径
    """
    # 复制模板文件
    shutil.copy2(template_excel_path, output_path)

    # 写入测试用例
    writer = ExcelWriter()
    writer.add_test_cases(test_cases)
    writer.write_to_excel(output_path)

    return output_path


def create_test_case_from_requirement(
    requirement_text: str,
    template_format: str = ""
) -> List[Dict[str, Any]]:
    """
    根据需求文本创建测试用例列表

    这是一个示例函数，实际的生成逻辑需要结合 AI 来完成

    Args:
        requirement_text: 需求文本
        template_format: 模板格式说明

    Returns:
        测试用例列表
    """
    # 这里应该调用 AI 来生成测试用例
    # 目前返回空列表，示意如何调用
    return []
