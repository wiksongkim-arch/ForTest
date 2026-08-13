"""
测试用例模板加载器 - 加载、结构化、缓存模板
"""
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from uuid import uuid4

from filelock import FileLock

from services.dingtalk_spreadsheet import DingTalkSpreadSheetMCPService


TEMPLATE_CACHE_PATH = "output/test_case_template.json"
TEMPLATE_CACHE_TTL = 24 * 60 * 60  # 24小时


class TemplateLoader:
    """测试用例模板加载器"""

    def __init__(
        self,
        template_url: str,
        spreadsheet_service: DingTalkSpreadSheetMCPService | None,
        cache_path: str = TEMPLATE_CACHE_PATH,
        local_template_path: str | Path | None = None,
    ):
        self.template_url = str(template_url or "").strip()
        self.spreadsheet_service = spreadsheet_service
        self.cache_path = cache_path
        self.local_template_path = (
            Path(local_template_path).expanduser().resolve()
            if local_template_path is not None
            else None
        )
        self._template_cache: Optional[Dict] = None

    def load_template(self, force_refresh: bool = False) -> Dict[str, Any]:
        """
        加载测试用例模板

        Args:
            force_refresh: 是否强制刷新缓存

        Returns:
            模板数据 {"field_specs": {...}, "components": {...}}
        """
        if not force_refresh and self._template_cache:
            return self._template_cache

        if self.template_url:
            cached = self._load_from_cache()
            if cached and not force_refresh:
                if self._is_cache_valid(cached):
                    self._template_cache = cached
                    return cached
            template_data = self._fetch_from_dingtalk()
            self._save_to_cache(template_data)
        else:
            # 本地副本允许用户自行调整，不复用在线缓存，也不记录绝对路径。
            template_data = self._load_from_local_template()
        self._template_cache = template_data
        return template_data

    def _load_from_cache(self) -> Optional[Dict[str, Any]]:
        """从本地缓存加载"""
        if not os.path.exists(self.cache_path):
            return None
        try:
            # 与并行任务的原子写配合，读取方不会看到尚未完成的 JSON。
            with FileLock(f"{self.cache_path}.lock"):
                with open(self.cache_path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except (json.JSONDecodeError, IOError):
            return None

    def _save_to_cache(self, data: Dict[str, Any]) -> None:
        """保存到本地缓存"""
        os.makedirs(os.path.dirname(self.cache_path), exist_ok=True)
        destination = Path(self.cache_path)
        temporary = destination.with_name(
            f".{destination.name}.{uuid4().hex}.tmp"
        )
        # 每个线程使用独立临时文件，完成刷新后再一次性替换目标文件。
        with FileLock(f"{self.cache_path}.lock"):
            try:
                with temporary.open("x", encoding="utf-8", newline="\n") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                    f.flush()
                    os.fsync(f.fileno())
                os.replace(temporary, destination)
            finally:
                temporary.unlink(missing_ok=True)

    def _is_cache_valid(self, cached: Dict[str, Any]) -> bool:
        """检查缓存是否有效"""
        if "updated_at" not in cached:
            return False
        try:
            updated_time = datetime.fromisoformat(cached["updated_at"])
            elapsed = (datetime.now() - updated_time).total_seconds()
            return elapsed < TEMPLATE_CACHE_TTL
        except (ValueError, TypeError):
            return False

    def _fetch_from_dingtalk(self) -> Dict[str, Any]:
        """从钉钉文档获取模板"""
        if self.spreadsheet_service is None:
            raise RuntimeError("钉钉表格 MCP 尚未配置")
        self._ensure_directory_exists()
        result = self.spreadsheet_service.get_document_content(
            self.template_url
        )

        if result.get("isError"):
            raise RuntimeError("获取模板文档失败")

        rows = self._parse_template_rows(result)
        return self._structure_template(rows)

    def _load_from_local_template(self) -> Dict[str, Any]:
        """从用户工作副本读取模板，打包母版永远不进入写路径。"""

        path = self.local_template_path
        if path is None or not path.is_file() or path.suffix.casefold() != ".xlsx":
            raise RuntimeError("本地用例模板不可用")
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if not workbook.worksheets:
                raise RuntimeError("本地用例模板没有工作表")
            sheet = workbook.worksheets[0]
            # 钉钉导出的 XLSX 可能把 dimension 错写成 A1:A1，读取前重算有效范围。
            reset_dimensions = getattr(sheet, "reset_dimensions", None)
            if callable(reset_dimensions):
                reset_dimensions()
            rows = [
                ["" if value is None else str(value) for value in values]
                for values in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
        return self._structure_template(rows)

    def _ensure_directory_exists(self) -> None:
        """确保目录存在"""
        directory = os.path.dirname(self.cache_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

    def _parse_template_rows(self, result: Dict[str, Any]) -> List[List[str]]:
        """
        解析模板文档的行数据

        Returns:
            二维数组，每行是一个列表
        """
        rows = []

        # 优先尝试从 content 解析（适用于 .adoc 文档）
        content = result.get("content", [])
        if content:
            for item in content:
                if item.get("type") == "text":
                    try:
                        data = json.loads(item.get("text", "{}"))
                        display_values = data.get("displayValues", [])
                        rows.extend(display_values)
                        break
                    except json.JSONDecodeError:
                        continue

        # 如果 content 为空，尝试从 texts 解析（适用于 .xlsx 表格）
        if not rows:
            texts = result.get("texts", [])
            for text in texts:
                text = text.strip()
                if not text or text.startswith("==="):
                    continue
                # 跳过表头描述行（包含"用例目录"等表头标记）
                # if "用例目录" in text or "用例等级" in text:
                #     continue
                # 使用 | 分隔列
                parts = [p.strip() for p in text.split("|")]
                # 过滤空列，但保留至少有一些内容的行
                non_empty = [p for p in parts if p]
                if non_empty:
                    rows.append(parts)

        return rows

    def _structure_template(self, rows: List[List[str]]) -> Dict[str, Any]:
        """
        结构化模板数据

        Args:
            rows: 原始行数据，第0行是标题，第1行是字段规范，第2行起是模板内容

        Returns:
            结构化模板 {"field_specs": {...}, "components": {...}}
        """
        if len(rows) < 2:
            return {"field_specs": {}, "components": {}}

        headers = rows[0] if rows else []
        # 从第1行获取表头并拼接第2行形成规范描述field_specs
        field_specs = self._parse_field_specs(rows[1] if len(rows) > 1 else [], headers)

        components: Dict[str, List[Dict]] = {}

        # 从第3行开始（第0行是标题，第1行是字段规范，第2行开始是模板内容）
        start_idx = 2
        if len(rows) > start_idx:
            for row in rows[start_idx:]:
                if not row or len(row) < 3:
                    continue

                component_name = row[2] if len(row) > 2 else ""  # 用例名称列
                if not component_name or not component_name.strip():
                    continue

                template_item = {
                    "用例步骤": row[4] if len(row) > 4 else "",  # 用例步骤列
                    "预期结果": row[5] if len(row) > 5 else ""   # 预期结果列
                }

                if component_name not in components:
                    components[component_name] = []
                components[component_name].append(template_item)

        return {
            "version": "1.0",
            "updated_at": datetime.now().isoformat(),
            "source_url": "",
            "field_specs": field_specs,
            "components": components
        }

    def _parse_field_specs(self, specs_row: List[str], headers: List[str]) -> Dict[str, str]:
        """
        解析字段规范说明

        Args:
            specs_row: 字段规范行数据
            headers: 表头行

        Returns:
            {"用例目录": "规范1", "用例等级": "规范2", ...}
        """
        field_specs = {}
        for i, header in enumerate(headers):
            header = header.strip()
            if header and i < len(specs_row):
                value = specs_row[i] if specs_row[i] else ""
                field_specs[header] = value.strip()
        return field_specs

    def get_component_names(self) -> List[str]:
        """获取所有组件名称"""
        template = self.load_template()
        return list(template.get("components", {}).keys())

    def get_components(self, names: List[str]) -> Dict[str, List[Dict]]:
        """获取指定名称的组件模板"""
        template = self.load_template()
        all_components = template.get("components", {})
        result = {}
        for name in names:
            if name in all_components:
                result[name] = all_components[name]
        return result

    def get_field_specs(self) -> Dict[str, str]:
        """获取字段规范说明"""
        template = self.load_template()
        return template.get("field_specs", {})


def create_template_loader(
    template_url: str,
    spreadsheet_service: DingTalkSpreadSheetMCPService | None,
    cache_path: str = TEMPLATE_CACHE_PATH,
    local_template_path: str | Path | None = None,
) -> TemplateLoader:
    """工厂函数：创建模板加载器"""
    return TemplateLoader(
        template_url=template_url,
        spreadsheet_service=spreadsheet_service,
        cache_path=cache_path,
        local_template_path=local_template_path,
    )
