"""ForTest 纯原生 Windows 桌面端入口。"""

from __future__ import annotations

import json
import os
import struct
import sys
import time
import traceback
from pathlib import Path
from tempfile import TemporaryDirectory

from windows_native.paths import (
    app_data_root,
    legacy_app_data_root,
    older_brand_app_data_root,
    prepare_runtime,
    previous_app_data_root,
    project_root,
    remove_user_data,
)
from windows_native.product import PRODUCT_NAME, PRODUCT_VERSION, eim_feature_enabled
from windows_native.process_policy import (
    install_hidden_subprocess_policy,
    is_policy_installed,
    terminate_owned_subprocesses,
)


def _argument_value(name: str) -> str | None:
    """读取简单的“参数 值”形式命令行参数。"""

    try:
        index = sys.argv.index(name)
    except ValueError:
        return None
    if index + 1 >= len(sys.argv):
        return None
    return sys.argv[index + 1]


DELETE_USER_DATA = "--delete-user-data" in sys.argv
EIM_ENABLED = eim_feature_enabled()


def _user_secret_names() -> set[str]:
    """返回卸载时应移除的固定及配置级密钥名称。"""

    from backend.settings.service import ai_configuration_secret_name

    names = {
        "document_mcp_url",
        "spreadsheet_mcp_url",
        "minimax_api_key",
        "openai_compatible_api_key",
        "codex_api_key",
        "jenkins_api_token",
    }
    for root in (
        app_data_root(),
        previous_app_data_root(),
        older_brand_app_data_root(),
        legacy_app_data_root(),
    ):
        try:
            payload = json.loads(
                (root / "data" / "settings.json").read_text(encoding="utf-8")
            )
            configurations = (payload.get("ai") or {}).get("configurations") or []
            for item in configurations:
                configuration_id = (
                    str(item.get("id") or "").strip()
                    if isinstance(item, dict)
                    else ""
                )
                if configuration_id:
                    names.add(ai_configuration_secret_name(configuration_id))
        except (OSError, TypeError, ValueError, json.JSONDecodeError):
            continue
    return names


def _delete_all_user_data() -> int:
    """由卸载器调用，删除统一数据目录及 Windows 凭据。"""

    try:
        from backend.settings.secrets import KeyringSecretStore

        secrets = KeyringSecretStore()
        names = _user_secret_names()
        # 配置文件损坏或已缺失时，仍从本产品凭据命名空间清理孤立动态密钥。
        names.update(secrets.list_names("ai_config:"))
        for name in names:
            secrets.delete(name)
        remove_user_data()
        return 0
    except Exception:
        return 1


def _run_backup_smoke_test() -> int:
    """在冻结环境中验证 openpyxl 导入和原子 XLSX 备份，不读取正式用户数据。"""

    diagnostics_value = _argument_value("--diagnostics-file")
    if not diagnostics_value:
        return 2
    payload: dict[str, object]
    try:
        from backend.ai.types import TEST_CASE_FIELDS
        from services.dingtalk_output import DingTalkOutputWriter
        from openpyxl import load_workbook

        with TemporaryDirectory(prefix="ForTest-backup-smoke-") as folder:
            root = Path(folder)
            writer = DingTalkOutputWriter(
                object(),
                object(),
                "https://example.test/template",
                "https://example.test/folder",
                lock_dir=root / "locks",
            )
            destination = root / "output" / "diagnostic.xlsx"
            writer._write_local_backup(
                destination,
                [{field: "诊断值" for field in TEST_CASE_FIELDS}],
            )
            workbook = load_workbook(destination, read_only=True)
            try:
                valid = (
                    destination.is_file()
                    and workbook.active.max_row == 2
                    and workbook.active.max_column == len(TEST_CASE_FIELDS)
                )
            finally:
                workbook.close()
            payload = {"success": valid, "error_type": None}
    except BaseException as exc:
        # 诊断只保存异常类型，避免第三方异常正文夹带本机路径或业务数据。
        payload = {"success": False, "error_type": type(exc).__name__}
    try:
        Path(diagnostics_value).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        return 2
    return 0 if payload["success"] else 1


# 删除模式不初始化 Qt、不创建单实例锁，也不会重建刚删除的数据目录。
if DELETE_USER_DATA:
    raise SystemExit(_delete_all_user_data())

BACKUP_SMOKE_TEST = "--backup-smoke-test" in sys.argv
if BACKUP_SMOKE_TEST:
    raise SystemExit(_run_backup_smoke_test())


# 打包自检必须与正式用户数据完全隔离，否则并行构建可能争用单实例锁，
# 还可能把正式任务误判为上次异常中断。一次性目录会在自检退出后删除。
SMOKE_TEST = "--smoke-test" in sys.argv
FULL_STARTUP_SMOKE = "--full-startup-smoke" in sys.argv
AUTOSTART = "--autostart" in sys.argv
DIAGNOSTIC_TEST = SMOKE_TEST or FULL_STARTUP_SMOKE
_DIAGNOSTICS_RUNTIME = (
    TemporaryDirectory(prefix="ForTest-smoke-") if DIAGNOSTIC_TEST else None
)

# 在加载任何业务模块前完成运行目录和子进程策略初始化。
DATA_ROOT = prepare_runtime(
    Path(_DIAGNOSTICS_RUNTIME.name) if _DIAGNOSTICS_RUNTIME is not None else None
)
install_hidden_subprocess_policy()

from PySide6.QtCore import Qt  # noqa: E402
from PySide6.QtGui import QIcon  # noqa: E402
from PySide6.QtWidgets import QApplication, QMessageBox  # noqa: E402


def _icon_path() -> Path:
    """兼容源码运行与 PyInstaller 解包目录。"""

    bundle = Path(getattr(sys, "_MEIPASS", project_root()))
    return bundle / "windows_native" / "assets" / "ForTester.ico"


def main() -> int:
    started_at = time.perf_counter()
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
    )
    app = QApplication(sys.argv)
    app.setApplicationName(PRODUCT_NAME)
    app.setApplicationVersion(PRODUCT_VERSION)
    app.setOrganizationName(PRODUCT_NAME)
    app.setStyle("Fusion")
    # 正式程序由托盘持有生命周期；诊断模式仍通过定时器显式退出。
    app.setQuitOnLastWindowClosed(False)
    icon_path = _icon_path()
    icon = QIcon(str(icon_path)) if icon_path.exists() else QIcon()
    app.setWindowIcon(icon)

    instance = None
    task_manager = None
    deployment_service = None
    splash = None
    acquired = False
    lifecycle = None
    lifecycle_heartbeat_timer = None
    exit_code = 1
    try:
        # 单实例检查走轻量模块，不再因导入主窗口而阻塞启动页首帧。
        from windows_native.single_instance import (
            SingleInstance,
            show_duplicate_instance_message,
        )

        if not DIAGNOSTIC_TEST:
            instance = SingleInstance()
            acquired = instance.acquire()
            if not acquired:
                # 覆盖升级时，旧安装版主进程可能仍持有兼容锁；只清理精确旧路径。
                from windows_native.legacy_cleanup import cleanup_previous_native_application

                if cleanup_previous_native_application():
                    acquired = instance.acquire()
            if not acquired:
                show_duplicate_instance_message()
                return 0

        from windows_native.lifecycle import (
            LifecycleDiagnostics,
            activate_lifecycle,
            current_exit_reason,
            lifecycle_event,
            request_application_exit,
        )

        launch_mode = (
            "diagnostic"
            if DIAGNOSTIC_TEST
            else "autostart"
            if AUTOSTART
            else "interactive"
        )
        lifecycle = LifecycleDiagnostics(
            DATA_ROOT,
            version=PRODUCT_VERSION,
            diagnostic_mode=DIAGNOSTIC_TEST,
        )
        lifecycle.start(launch_mode=launch_mode)
        activate_lifecycle(lifecycle)

        # 覆盖 Qt 自身退出、末窗口关闭和 Windows 注销/关机三类入口。
        app.aboutToQuit.connect(
            lambda: lifecycle_event(
                "qt_about_to_quit",
                reason=current_exit_reason() or "not_recorded",
            )
        )
        app.lastWindowClosed.connect(
            lambda: lifecycle_event("qt_last_window_closed")
        )
        commit_data_request = getattr(app, "commitDataRequest", None)
        if commit_data_request is not None:
            commit_data_request.connect(
                lambda _manager: request_application_exit("windows_session_end")
            )
        save_state_request = getattr(app, "saveStateRequest", None)
        if save_state_request is not None:
            save_state_request.connect(
                lambda _manager: lifecycle_event("windows_session_state_save")
            )

        from windows_native.ui.startup_splash import (
            StartupLoader,
            StartupSplash,
            wait_for_startup,
        )

        # 先向用户交付一个真实可重绘的窗口，再继续加载页面和本地服务。
        splash = StartupSplash(icon, PRODUCT_NAME, f"v{PRODUCT_VERSION}")
        splash.set_stage("", 6)
        splash.center_on_screen()
        if not AUTOSTART:
            splash.show()
        app.processEvents()
        splash_first_paint_seconds = time.perf_counter() - started_at

        from windows_native.desktop_preferences import DesktopPreferences
        from windows_native.i18n import set_language, tr
        from windows_native.ui.theme import ThemeManager

        preferences = DesktopPreferences(DATA_ROOT)
        set_language(preferences.get_language())
        theme_manager = ThemeManager(app, preferences)
        theme_manager.apply()
        splash.apply_theme(theme_manager.effective_mode())
        splash.set_stage(tr("正在准备本地数据…"), 28)
        app.processEvents()

        # 惰性门面仍负责线程安全，但正式启动会在启动页内主动解析完成。
        from windows_native.lazy_service import (
            LazyJenkinsDeploymentService,
            LazyNativeService,
        )

        service = LazyNativeService(DATA_ROOT)
        deployment_service = LazyJenkinsDeploymentService(DATA_ROOT)
        startup_snapshot = None
        startup_preload_complete = False
        backend_ready_before_main = False
        deployment_ready_before_main = False

        def create_task_manager():
            # 任务文件解析与异常中断修复也属于启动加载，必须留在启动页阶段。
            from windows_native.task_manager import TaskManager

            return TaskManager(service, DATA_ROOT, preferences)

        def cleanup_legacy() -> None:
            """清理旧 WebView 进程；失败不影响当前原生客户端启动。"""

            try:
                from windows_native.legacy_cleanup import (
                    cleanup_legacy_desktop_processes,
                )

                cleanup_legacy_desktop_processes()
            except Exception:
                return

        if SMOKE_TEST:
            splash.set_stage(tr("正在检查运行环境…"), 44)
            task_manager = create_task_manager()
            deployment_ready_before_main = True
        else:
            loader = StartupLoader(
                service,
                deployment_service,
                create_task_manager,
                cleanup=cleanup_legacy,
            )
            if not wait_for_startup(
                loader,
                splash,
                translate=tr,
                timeout_seconds=45.0,
            ):
                raise TimeoutError("本地业务初始化超过 45 秒")
            startup_snapshot = loader.result()
            task_manager = startup_snapshot.task_manager
            startup_preload_complete = True
            backend_ready_before_main = "backend.api.routes" in sys.modules
            deployment_ready_before_main = True

        splash.set_stage(tr("正在构建主界面…"), 90)
        app.processEvents()
        from windows_native.ui.main_window import MainWindow

        window = MainWindow(
            service,
            icon,
            task_manager,
            theme_manager,
            deployment_service,
            onboarding_enabled=not DIAGNOSTIC_TEST and not AUTOSTART,
            background_refresh_enabled=not DIAGNOSTIC_TEST,
            startup_preloaded=startup_preload_complete,
            tray_enabled=not DIAGNOSTIC_TEST,
            eim_enabled=EIM_ENABLED,
        )
        if startup_snapshot is not None:
            window.apply_startup_snapshot(startup_snapshot)
        splash.set_stage(tr("正在完成启动准备…"), 96)
        app.processEvents()
        window.apply_adaptive_geometry()
        splash.set_stage(tr("准备就绪"), 100)
        if AUTOSTART and not DIAGNOSTIC_TEST:
            # 开机启动只建立托盘和后台服务，不抢焦点或弹出向导。
            splash.close()
            window.hide()
        else:
            splash.finish(window)
        app.processEvents()
        first_paint_seconds = time.perf_counter() - started_at
        lifecycle_event(
            "main_window_ready",
            startup_seconds=round(first_paint_seconds, 3),
            visible=bool(window.isVisible()),
            tray_usable=bool(window._tray_is_usable()),
            close_behavior=window._preferred_close_behavior(),
        )

        # 正式入口此时只安排向导判断和外部网络刷新，所有本地加载均已完成。
        if not SMOKE_TEST:
            window.start_background_services()

        from PySide6.QtCore import QThreadPool, QTimer

        def sample_lifecycle_heartbeat() -> None:
            state = app.applicationState()
            lifecycle.heartbeat(
                application_state=str(getattr(state, "name", state)),
                window_visible=bool(window.isVisible()),
                window_minimized=bool(window.isMinimized()),
                tray_usable=bool(window._tray_is_usable()),
            )

        sample_lifecycle_heartbeat()
        lifecycle_heartbeat_timer = QTimer(app)
        lifecycle_heartbeat_timer.setInterval(30_000)
        lifecycle_heartbeat_timer.timeout.connect(sample_lifecycle_heartbeat)
        lifecycle_heartbeat_timer.start()
        lifecycle_event("lifecycle_heartbeat_started", interval_seconds=30)

        diagnostics_path = _argument_value("--diagnostics-file")
        post_show_heartbeat = {
            "ticks": 0,
            "threadpool_peak": 0,
            "last_at": None,
            "max_gap_seconds": 0.0,
        }
        diagnostics_timer = None
        if diagnostics_path:
            diagnostics_timer = QTimer(app)
            diagnostics_timer.setInterval(40)

            def sample_post_show() -> None:
                # 主窗口后的独立心跳用于验收：若再发生重加载，计数会明显停滞。
                now = time.monotonic()
                last_at = post_show_heartbeat["last_at"]
                if isinstance(last_at, float):
                    post_show_heartbeat["max_gap_seconds"] = max(
                        post_show_heartbeat["max_gap_seconds"],
                        now - last_at,
                    )
                post_show_heartbeat["last_at"] = now
                post_show_heartbeat["ticks"] += 1
                post_show_heartbeat["threadpool_peak"] = max(
                    post_show_heartbeat["threadpool_peak"],
                    QThreadPool.globalInstance().activeThreadCount(),
                )

            diagnostics_timer.timeout.connect(sample_post_show)
            diagnostics_timer.start()
        if DIAGNOSTIC_TEST:
            def finish_diagnostics() -> None:
                request_application_exit("diagnostic_timer")
                app.quit()

            QTimer.singleShot(
                1500 if FULL_STARTUP_SMOKE else 900,
                finish_diagnostics,
            )

        result = app.exec()
        exit_code = int(result)
        lifecycle_event("qt_event_loop_returned", exit_code=exit_code)
        if current_exit_reason() is None:
            request_application_exit("qt_event_loop_returned_without_request")
        if diagnostics_timer is not None:
            diagnostics_timer.stop()
        if diagnostics_path:
            geometry = window.geometry()
            payload = {
                "architecture_bits": struct.calcsize("P") * 8,
                "native_qt": True,
                "web_server_started": False,
                "diagnostics_isolated": DIAGNOSTIC_TEST,
                "full_startup_smoke": FULL_STARTUP_SMOKE,
                "backend_runtime_loaded": "backend.api.routes" in sys.modules,
                "codex_runtime_loaded": any(
                    module == "openai_codex"
                    or module.startswith("openai_codex.")
                    or module == "codex_cli_bin"
                    or module.startswith("codex_cli_bin.")
                    for module in sys.modules
                ),
                "product": PRODUCT_NAME,
                "version": PRODUCT_VERSION,
                "theme_mode": theme_manager.mode,
                "subprocess_policy": is_policy_installed(),
                "startup_seconds": round(first_paint_seconds, 3),
                "first_paint_seconds": round(first_paint_seconds, 3),
                "splash_first_paint_seconds": round(splash_first_paint_seconds, 3),
                "startup_heartbeat_ticks": splash.heartbeat_count,
                "startup_max_heartbeat_gap_seconds": round(
                    splash.max_heartbeat_gap_seconds,
                    3,
                ),
                "startup_preload_complete": startup_preload_complete,
                "startup_snapshot_applied": bool(
                    getattr(window, "_startup_snapshot_applied", False)
                ),
                "backend_ready_before_main": backend_ready_before_main,
                "deployment_ready_before_main": deployment_ready_before_main,
                "post_show_heartbeat_ticks": post_show_heartbeat["ticks"],
                "post_show_max_heartbeat_gap_seconds": round(
                    post_show_heartbeat["max_gap_seconds"],
                    3,
                ),
                "post_show_threadpool_peak": post_show_heartbeat["threadpool_peak"],
                "window": {
                    "width": geometry.width(),
                    "height": geometry.height(),
                    "x": geometry.x(),
                    "y": geometry.y(),
                },
                "data_root": str(DATA_ROOT),
            }
            Path(diagnostics_path).write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        return result
    except BaseException as exc:
        exit_code = 1
        if lifecycle is not None:
            lifecycle.record_exception(
                "main_exception",
                type(exc),
                exc,
                exc.__traceback__,
                context="startup" if "window" not in locals() else "event_loop",
            )
            lifecycle.request_exit("main_exception")
        if splash is not None:
            splash.close()
        log_path = DATA_ROOT / "logs" / "startup.log"
        log_path.write_text(
            "".join(traceback.format_exception(exc)),
            encoding="utf-8",
        )
        if DIAGNOSTIC_TEST:
            # 无界面构建自检不能停在模态错误框；完整堆栈交给构建日志判定。
            traceback.print_exception(exc, file=sys.stderr)
        else:
            QMessageBox.critical(
                None,
                f"{PRODUCT_NAME} 启动失败",
                f"程序启动失败，诊断日志已保存到：\n{log_path}",
            )
        return 1
    finally:
        if lifecycle_heartbeat_timer is not None:
            lifecycle_heartbeat_timer.stop()
        if lifecycle is not None:
            lifecycle.record("shutdown_cleanup_started")
        if "service" in locals() and service is not None:
            service.stop_loaded_services()
        if task_manager is not None:
            task_manager.stop()
        if deployment_service is not None:
            deployment_service.stop()
        terminate_owned_subprocesses()
        if lifecycle is not None:
            lifecycle.record("shutdown_cleanup_finished")
            lifecycle.finish(
                exit_code=exit_code,
                fallback_reason="main_returned",
            )
        # 活动会话标记必须先于单实例锁释放，避免快速重启把正常收尾误判为崩溃。
        if instance is not None and acquired:
            instance.release()
        if _DIAGNOSTICS_RUNTIME is not None:
            # Windows 不能删除进程当前工作目录，先回到其父目录再回收。
            try:
                os.chdir(Path(_DIAGNOSTICS_RUNTIME.name).parent)
                _DIAGNOSTICS_RUNTIME.cleanup()
            except OSError:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
